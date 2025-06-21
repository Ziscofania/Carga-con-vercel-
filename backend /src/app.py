from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import sqlite3
import os
import openpyxl
from io import BytesIO

app = Flask(__name__)
CORS(app)

# Configuración de la base de datos
DATABASE = 'inventory.db'

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    if not os.path.exists(DATABASE):
        conn = get_db_connection()
        conn.execute('''
            CREATE TABLE products (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                category TEXT,
                notes TEXT
            )
        ''')
        conn.commit()
        conn.close()

# Rutas de la API
@app.route('/api/products', methods=['GET'])
def get_products():
    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products').fetchall()
    conn.close()
    return jsonify([dict(product) for product in products])

@app.route('/api/products', methods=['POST'])
def add_product():
    data = request.get_json()
    
    conn = get_db_connection()
    conn.execute(
        'INSERT INTO products (id, name, quantity, category, notes) VALUES (?, ?, ?, ?, ?)',
        (data['id'], data['name'], data['quantity'], data.get('category', ''), data.get('notes', ''))
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/products/<id>', methods=['PUT'])
def update_product(id):
    data = request.get_json()
    
    conn = get_db_connection()
    conn.execute(
        'UPDATE products SET name = ?, quantity = ?, category = ?, notes = ? WHERE id = ?',
        (data['name'], data['quantity'], data.get('category', ''), data.get('notes', ''), id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/products/<id>', methods=['DELETE'])
def delete_product(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM products WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/products/export', methods=['GET'])
def export_products():
    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products').fetchall()
    conn.close()
    
    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Encabezados
    ws.append(['ID', 'Nombre', 'Cantidad', 'Categoría', 'Notas'])
    
    # Datos
    for product in products:
        ws.append([product['id'], product['name'], product['quantity'], 
                  product['category'], product['notes']])
    
    # Guardar en un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name='inventario.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/products/import', methods=['POST'])
def import_products():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        conn = get_db_connection()
        
        # Leer filas (asumiendo que la primera fila son encabezados)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Asumimos el orden: ID, Nombre, Cantidad, Categoría, Notas
            product_id = row[0] or str(int(time.time() * 1000))  # Generar ID si no existe
            name = row[1] or 'Producto sin nombre'
            quantity = int(row[2] or 0)
            category = row[3] or ''
            notes = row[4] or ''
            
            # Verificar si el producto ya existe
            existing = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
            
            if existing:
                # Actualizar cantidad
                new_quantity = existing['quantity'] + quantity
                conn.execute(
                    'UPDATE products SET quantity = ? WHERE id = ?',
                    (new_quantity, product_id)
                )
            else:
                # Insertar nuevo producto
                conn.execute(
                    'INSERT INTO products (id, name, quantity, category, notes) VALUES (?, ?, ?, ?, ?)',
                    (product_id, name, quantity, category, notes)
                )
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True)